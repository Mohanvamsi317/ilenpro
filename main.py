import mysql.connector                   
import requests                          
from collections import Counter
from pathlib import Path                 
import pandas as pd                      
import openpyxl                          
import logging                           
from tqdm import tqdm                    



# for logging information
logging.basicConfig(filename='logfile.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s',filemode='w')
#to establish the database connection
class DatabaseConnection:
    def __init__(self, host, user, password, database):
        self.host = host
        self.user = user
        self.password = password
        self.database = database
        self.conn = None


    def connect(self):
        try:
            self.conn = mysql.connector.connect(
                host=self.host,
                user=self.user,
                password=self.password,
                database=self.database
            )
            print("Connection established successfully!")
            logging.info("Coonection established successfully")
        except mysql.connector.Error as err:
            print(f"Error: {err}")
            logging.exception("connection failed")
    # close the database connection
    def close(self):
        if self.conn:
            self.conn.close()
            logging.info("connection closed")

class Connect:
    def __init__(self, conn):
        self.conn = conn
        self.cursor = self.conn.cursor()
        self.link = []
        self.path = []
        self.attachment_dict = {}
        self.list_of_duplicate_byparts = []
        self.m_list=[]
        self.count_of_total_links=0
        
    # Separate the downloadable links and paths from the database and find thier count.
    def process_attachments(self):
        #fetch the attachment table in the database
        query = "SELECT * FROM attachments"
        self.cursor.execute(query)
        self.result = self.cursor.fetchall()
        self.column_index = self.cursor.column_names.index('attachment_filename')
        self.attachments_type = self.cursor.column_names.index('approval_type')
        
        
        query_by_part = "SELECT * FROM attachments WHERE approval_type = 'By part'"
        self.cursor.execute(query_by_part)
        self.result_by_part = self.cursor.fetchall()
        self.result_attachmentname = self.cursor.column_names.index("attachment_filename")
        self.result_component_id=self.cursor.column_names.index("component_id")
        
        #fetch the manfacturer table
        query1 = "SELECT * FROM manufacturers"
        self.cursor.execute(query1)
        self.manufacture_results = self.cursor.fetchall()
        #column_names.index is a method to access the index of the particular colum as the list item cant be accessed by its 
        self.manufacture_name_manufacture = self.cursor.column_names.index('name')
        self.manufucture_index = self.cursor.column_names.index('normalized_name')
        self.manufacturer_id_index = self.cursor.column_names.index('id')

        # Query to execute the master_components table
        query2 = "SELECT * FROM master_components"
        self.cursor.execute(query2)
        self.component_results = self.cursor.fetchall()
        self.manufacture_part_in_master =self.cursor.column_names.index('manufacturer_partnumber')
        self.component_id_component = self.cursor.column_names.index('id')
        self.manufacture_name_component =self.cursor.column_names.index('manufacturer_name')
        count=0
         # if filename starts with http or https the it was a downloadable else it is a path
        for row in self.result:
            if row[self.column_index].startswith("https://") or row[self.column_index].startswith("http://"):
                self.link.append(row[self.column_index])
                
            else:
                self.path.append(row[self.column_index])
            count=count+1
        self.count_of_total_links=count
        self.count_of_valid_links = len(self.link)
        self.path_for_excel=Path(input("enter the path where you want to store the excel files:"))


    # function to download the each file based on thier category
    def download_all_file(self):
        print("entered download_all_files")
        #user defined input to make sure where the particular file needed to be stored in local system
        pat = input("Enter the directory path where you wanty to store the files:")
        #ask the  user weather he can replace the existing one or not
        replace_choice = input("Do you want to replace existing files? (yes/no): ").strip().lower()
        if replace_choice not in ("yes", "no", 'y', 'n'):
            logging.debug("invalid choice for replacing the file")
            # print("Invalid choice. Exiting...")
            return
        try:    
            #to display the processing bar
            total_files = self.count_of_total_links
            pbar= tqdm(total=total_files, desc="Downloading files")
           #download the by part files that are mapped to multiple manufacturers
            for row in self.list_of_duplicate_byparts:
                q = "select *  from attachments where attachment_filename=%s and approval_type='By part' group by manufacturer_id "
                self.cursor.execute(q, (row,))
                res = self.cursor.fetchall()
                a_id = self.cursor.column_names.index("id")
                afn = self.cursor.column_names.index("attachment_filename")
                a_mid = self.cursor.column_names.index("manufacturer_id")
                if len(res) > 1:
                    for m in res:
                        id_for_attachment = m[a_id]
                        dem = m[afn]
                        attachment_mid = m[a_mid]
                        que = "select * from attachments where attachment_filename=%s and manufacturer_id=%s and approval_type='By part'"
                        self.cursor.execute(que, (dem, attachment_mid))
                        res_attachment = self.cursor.fetchall()
                        #files that are associated to multiple manufactures but they have similar approval type as by part those file reside in the compilance directory located inside the root directory
                        for i in res_attachment:
                            dir_name_multiple = "compilance"
                            local_directory_multiple = Path(pat) / dir_name_multiple
                            na = Path(dem)
                            last_name = na.name
                            file_name = Path(local_directory_multiple) / last_name
                            logging.info("The file is organized as bypart, but it was mapped to multiple manufacturers.")


                            logging.info(f"The File is:{dem}")
                            if local_directory_multiple.exists():
                                logging.debug(f"This folder {local_directory_multiple} already exists")
                                #print("already present multiple fiolder")
                                if (replace_choice == "yes" or replace_choice == "y"):
                                    self.download_file(dem, local_directory_multiple, id_for_attachment)
                                    logging.info("File has been overridden")
                                else:
                                    logging.info(f"{file_name}")
                                    logging.info("File already present in the specified location")
                                pbar.update(1)
                            else:
                                local_directory_multiple.mkdir()
                                logging.debug(f"Directory for multiple files was created {local_directory_multiple}")
                                #print("Directory for multiple files was created :", local_directory_multiple)
                                logging.info(f"{file_name}")
                                self.download_file(dem, local_directory_multiple, id_for_attachment)
                                
                                pbar.update(1)
                            self.m_list.append(dem)
            
            # to create the directory for the each manufacturer with the normalized_name
            for row in self.manufacture_results:
                manufacturer_id = row[self.manufacturer_id_index]
                directory_name = row[self.manufucture_index]
                manu_name = row[self.manufacture_name_manufacture]
                local_directory = Path(pat) / directory_name
                if local_directory.exists():
                    logging.debug("already exists:%s", local_directory)
                else:
                    local_directory.mkdir()
                    logging.debug("Directory created:%s", local_directory)
                #create the directory named as compilance inside the normalized_name for each manufacturer
                directory_name_comp = 'compilance'
                local_directory_compilance = local_directory / directory_name_comp

                if local_directory_compilance.exists():
                    logging.debug(f"already present compilance directory{local_directory_compilance}")
                else:
                    local_directory_compilance.mkdir()
                    logging.debug("Compilance directory created: %s", local_directory_compilance)

                #create the directory inside the compilance folder with manufacturer_partnumber
                query2 = "SELECT * FROM master_components WHERE manufacturer_name = %s"
                self.cursor.execute(query2, (manu_name,))
                component_results = self.cursor.fetchall()

                manufacture_part_in_master = self.cursor.column_names.index('manufacturer_partnumber')
                component_id_component = self.cursor.column_names.index('id')
              
                manufacture_name_component = self.cursor.column_names.index('manufacturer_name')

                for component_row in component_results:
                    if component_row[manufacture_name_component] == row[self.manufacture_name_manufacture]:
                        component_id = component_row[component_id_component]
                        directory_name1 = component_row[manufacture_part_in_master]
                        local_directory2 = local_directory_compilance / directory_name1

                        if local_directory2.exists():
                            logging.debug("There is already a folder for the entered component part")
                        else:
                            local_directory2.mkdir(parents=True, exist_ok=True)
                            logging.debug("Subdirectory created: %s", local_directory2)

                        #download the files which are clean and without duplicates
                        #here the elements in the list gets seperated at first the total of some spaces were seperated by , then the places were replaced with the iten in the list
                        attachment_placeholders = ', '.join(['%s'] * len(self.list_of_duplicate_byparts))
                        query3 = f"SELECT * FROM attachments WHERE manufacturer_id = %s AND component_id = %s AND attachment_filename NOT IN ({attachment_placeholders})"
                        query_params = [manufacturer_id, component_id] + self.list_of_duplicate_byparts
                        self.cursor.execute(query3, query_params)
                        attachment_results = self.cursor.fetchall()
                        id_of_index = self.cursor.column_names.index('id')
                        att_res_lin = self.cursor.column_names.index("attachment_filename") 
                        for attachment_row in attachment_results:
                            file_url = attachment_row[att_res_lin]
                            file_type = attachment_row[self.attachments_type]
                            id_of_a = attachment_row[id_of_index]
                            
                            if file_type == "partseries":
                                directory_name_partseries = 'partseries'
                                local_directory_partseries = local_directory_compilance / directory_name_partseries
                                logging.info("Its a Partseries file")
                                logging.info(f"File is: {file_url}")
                                if local_directory_partseries.exists():
                                    logging.debug("already exists : %s", local_directory_partseries)
                                    na = Path(file_url)
                                    last_name = na.name
                                    file_name = Path(local_directory_partseries) / last_name
                                    if (file_name.exists()):
                                        logging.debug("already present:")
                                        if (replace_choice == 'y' or replace_choice == 'yes'):
                                            self.download_file(file_url, local_directory_partseries, id_of_a)
                                            logging.info("File has been overridden")
                                        else:
                                            logging.info(f"File is Already present in the {local_directory_partseries}")
                                        pbar.update(1)
                                else:
                                    local_directory_partseries.mkdir()
                                    logging.debug("Partseries directory created: %s", local_directory_partseries)
                                    self.download_file(file_url, local_directory_partseries, id_of_a)
                                    logging.info("File has been downloaded and saved")
                                    pbar.update(1)
                            elif file_type == "blanket":
                                directory_name_blanket = 'blanket'
                                local_directory_blanket = local_directory_compilance / directory_name_blanket
                                logging.info("Its a Blanket file")
                                logging.info(f"File is: {file_url}")
                                pbar.update(1)
                                if local_directory_blanket.exists():
                                    logging.debug("directory for the blanket already exist: %s", local_directory_blanket)
                                    na = Path(file_url)
                                    last_name = na.name
                                    file_name = Path(local_directory_blanket) / last_name
                                    if (file_name.exists()):
                                        logging.debug("Blanket file already present")
                                        if (replace_choice == 'y' or replace_choice == 'yes'):
                                            self.download_file(file_url, local_directory_multiple, id_of_a)
                                            logging.info("File has been overridden")
                                        else:
                                            logging.info(f"File is: {file_name}")
                                            logging.info(f"File already present in the {local_directory_blanket}")
                                        pbar.update(1)
                                else:
                                    local_directory_blanket.mkdir()
                                    logging.debug("Directory for the Blanket files has been created:%s", local_directory_blanket)
                                    self.download_file(file_url, local_directory_blanket, id_of_a)
                                    logging.info("Blanket file is downloaded and saved")
                                    pbar.update(1)

                            else:
                                na = Path(file_url)
                                last_name = na.name
                                file_name = Path(local_directory2) / last_name
                                logging.info("The file is a 'ByPart' file and it doesn't have mappings to any other manufacturers or components.")
                                logging.info(f"File is: {file_url}")
                                if (file_name.exists()):
                                    logging.debug("file already present in:%s", local_directory2)
                                    if (replace_choice == 'y' or replace_choice == 'yes'):
                                        self.download_file(file_url, local_directory2, id_of_a)
                                        logging.info("File has been overridden")
                                    else:
                                        logging.info(f"File is Already present in the {local_directory2}")
                                    pbar.update(1)
                                else:
                                    self.download_file(file_url, local_directory2, id_of_a)
                                    logging.info("Bypart file is downloaded and saved")
                                    pbar.update(1)
             # down the files which are  byparts and blanket or bypart and partseries:
            for k in self.list_of_duplicate_byparts:
                query_for_blank_part = "select * from attachments where attachment_filename=%s and approval_type !='By part'"
                self.cursor.execute(query_for_blank_part, (k,))

                res = self.cursor.fetchall()
                index_man = self.cursor.column_names.index("manufacturer_id")
                index_id = self.cursor.column_names.index("id")
                index_approval = self.cursor.column_names.index("approval_type")
                for i in res:
                    manu_id = i[index_man]
                    id_for_id = i[index_id]
                    app_type = i[index_approval]
                    query_for_manu = "select * from manufacturers where id=%s"
                    self.cursor.execute(query_for_manu, (manu_id,))
                    res_query_manu = self.cursor.fetchall()
                    manu_name_id = self.cursor.column_names.index("normalized_name")
                    for j in res_query_manu:
                        name = j[manu_name_id]
                        if (app_type == "partseries"):
                            logging.info("Its a partseries file")
                            local_dir_bypart_partseries = Path(pat) / name / "compilance" / "partseries"
                            na = Path(k)
                            last_name = na.name
                            file_name = Path(local_dir_bypart_partseries) / last_name
                            logging.info("this file is mapped to both bypart and partseries")
                            if (local_dir_bypart_partseries.exists()):
                                if (replace_choice == "y" or replace_choice == "yes"):
                                    self.download_file(k, local_dir_bypart_partseries, id_for_id)
                                    
                                else:
                                    logging.info(f"File {file_name}")
                                    logging.info(f"File is already present in {local_dir_bypart_partseries}")
                                pbar.update(1)
                            else:
                                local_dir_bypart_partseries.mkdir()
                                self.download_file(k, local_dir_bypart_partseries, id_for_id)
                                logging.info("This File is mapped to partseries and bypartfile")
                                pbar.update(1)
                        if (app_type == "blanket"):
                            logging.info("blanket")
                            local_dir_bypart_blanket = Path(pat) / name / "compilance" / "blanket"
                            na = Path(k)
                            last_name = na.name
                            file_name = Path(local_dir_bypart_blanket) / last_name
                            logging.info("This File is mapped to both bypart and blanket")
                            if (local_dir_bypart_blanket.exists()):
                                if (replace_choice == "y" or replace_choice == "yes"):
                                    self.download_file(k, local_dir_bypart_blanket, id_for_id)
                                    
                                else:
                                    logging.info(f"file is {file_name}")
                                    logging.info("file is already present")
                                pbar.update(1)
                            else:
                                local_dir_bypart_blanket.mkdir()
                                self.download_file(k, local_dir_bypart_blanket, id_for_id)
                                logging.info("blanket and bypartfile")
                                pbar.update(1)
            # a list of files that have been mapped to several manufacturers of approval type as bypart
            self.m_list = set(self.m_list)
            logging.debug("these are the files that are mapped to the same different manufacturers and these are the bypart files:%s",self.m_list)
            
            for k in self.m_list:
                self.list_of_duplicate_byparts.remove(k)
            # download the files which are mapped to more than one manufacturer and same approval tye as the by part
            for i in self.list_of_duplicate_byparts:
                qp = "SELECT * FROM attachments WHERE approval_type='By part' AND attachment_filename=%s"
                self.cursor.execute(qp, (i,))
                m_id = self.cursor.column_names.index("manufacturer_id")
                ans = self.cursor.fetchall()
                id_in_index = self.cursor.column_names.index("id")
                for j in ans:
                    id_in_attachments = j[id_in_index]
                    aq = "SELECT * FROM manufacturers WHERE id=%s"
                    self.cursor.execute(aq, (j[m_id],))
                    ans_in_j = self.cursor.fetchall()
                    n = self.cursor.column_names.index("normalized_name")
                    for k in ans_in_j:
                        n_name = k[n]
                        dir = Path(pat) / n_name / "compilance" / "multiplebyparts"
                        if dir.exists():
                            logging.debug("already multiple exists")
                            if (replace_choice == 'y' or replace_choice == 'yes'):
                                self.download_file(i, dir, id_in_attachments)
                                
                            else:
                                logging.info(f"file already present in the {dir}")
                            pbar.update(1)
                        else:
                            dir.mkdir(parents=True, exist_ok=True)
                            logging.debug("created folder for the multiple byparts but mapped to same manufacturer")
                            self.download_file(i, dir, id_in_attachments)
                            logging.info(f"file for multiple byparts has been downloaded and saved in the{dem}")
                            pbar.update(1)
            

            
            pbar.close()

        except Exception as e:
            logging.error(e)
            print(f"Error: {e}")



    # function to store each file in the location after downloading it from the URLs       
    def download_file(self, file_url, local_directory, id_of_attachments):
        try:
            response = requests.get(file_url)
            response.raise_for_status()  # Raise an exception for bad status codes
            #from the url the last most positon is named as the file name and saved with that name
            filename = local_directory / Path(file_url.split('?')[0]).name.strip()
            logging.info(id_of_attachments)
            
            with open(filename, 'wb') as local_file:
                local_file.write(response.content)
                logging.info("File downloaded and saved in: %s", filename)
                        
            # Manipulate file path to get the relative path from 'atom_mfr_docs'
            pa = Path(local_directory) / filename
            part = pa.parts
            index = part.index("atom_mfr_docs")
            list_of_parts = list(part[index:])
            finalpath = "\\".join(list_of_parts)
            
            # Update the database record with the path of the downloaded file
            query_to_update = "UPDATE attachments SET notes=%s WHERE id=%s"
            self.cursor.execute(query_to_update, (finalpath, id_of_attachments))
            logging.info("File path updated in the database: %s", finalpath)
            #print("file path updated in the database:",finalpath)
            
            # Commit the transaction
            self.conn.commit()
            
            
            return True  # Return True to indicate successful download and database update

        except Exception as e:
            logging.error("An exception occurred while downloading the file: %s", e)
            return False  
    # function to determine the overall number of links and thier approval type links (bypart, partseries, and blanket)

    def count_of_linktypes(self):
        dict_count_links = {}
        for link_item in self.link:
            count_bypart = 0
            count_partseries = 0
            count_blanket = 0
            for row in self.result:
                if row[self.column_index] == link_item:
                    if row[self.attachments_type] == 'By part': 
                        count_bypart += 1
                    elif row[self.attachments_type] == 'partseries':
                        count_partseries += 1
                    elif row[self.attachments_type] == 'blanket':
                        count_blanket += 1
                total_count = count_bypart + count_partseries + count_blanket
            dict_count_links[link_item] = {'By part': count_bypart, 'Partseries': count_partseries, 'Blanket': count_blanket , 'count': total_count}
        # Convert the dictionary to DataFrame
        df_link_count = pd.DataFrame.from_dict(dict_count_links, orient='index').reset_index()
        df_link_count.columns = ['Link', 'By part', 'Partseries', 'Blanket', 'count']

        # Write the DataFrame to an Excel file
        files_linktypes = self.path_for_excel / 'count_of_linktypes.xlsx'
        with pd.ExcelWriter(files_linktypes) as writer:
            df_link_count.to_excel(writer, index=False)

    # function to find out the total no of duplicate byparts in the database
    def duplicate_byparts(self):
        self.attachment_dict = {}
        self.count_of_duplicatebyparts=0
        for link in self.link:
            for row in self.result:
                if row[self.attachments_type] == 'By part' and row[self.column_index]==link:
                    attachment_filename=row[self.column_index]
                    query_component_id = "SELECT component_id FROM attachments WHERE attachment_filename = %s AND approval_type = 'By part'"
                    self.cursor.execute(query_component_id, (attachment_filename,))
                    result_component_ids = self.cursor.fetchall()
                
                     # Extract component IDs
                    component_ids = [res[0] for res in result_component_ids]

                # Check if there are duplicate component IDs
                    if len(set(component_ids)) > 1:
                        self.attachment_dict[attachment_filename] = component_ids
        for attachment,component_id in self.attachment_dict.items():
            self.attachment_dict[attachment]=set(component_id)
        self.count_of_duplicatebyparts=len(self.attachment_dict)

        if self.attachment_dict:
            df_attachment_dict = pd.DataFrame(self.attachment_dict.items(), columns=['Attachment', 'Component IDs'])
           
            df_total_links = pd.DataFrame([['total links', self.count_of_valid_links]],
                                          columns=['Attachment', 'Count'])
            count_of_uniquelinks = len(set(self.link))
            
            # write the data content to the excel sheet and store the data in the folder
            filename_duplicate_bypart=self.path_for_excel / 'list_of_duplicate_byparts.xlsx'
            with pd.ExcelWriter(filename_duplicate_bypart) as writer:
                df_attachment_dict.to_excel(writer, index=False)

            filename_total = self.path_for_excel / 'final_report.xlsx'
            if filename_total.exists():
                self.clear_excel_file(filename_total)
                wb = openpyxl.load_workbook(filename_total)
                ws = wb.active

                ws.append(['count of total links', len(self.link)])
                ws.append(['count of unique links', count_of_uniquelinks])
                ws.append(['count of duplicate by parts', self.count_of_duplicatebyparts])
                wb.save(filename_total)

            else:
                with pd.ExcelWriter(filename_total) as writer:
                    df_total_links.to_excel(writer, index=False)
                wb = openpyxl.load_workbook(filename_total)
                ws = wb.active
                ws.append(['count of unique links ', count_of_uniquelinks])
                ws.append(['count of duplicate by parts', self.count_of_duplicatebyparts])
                wb.save(filename_total)
            self.count_of_total_links=count_of_uniquelinks+self.count_of_duplicatebyparts    
                
    def partseries(self):
        self.count_of_partseries_bypart=0
        #print("partseries")
        logging.info("Part series")

        
        # Query to select attachments with approval type 'partseries'
        query_part_series = "SELECT * FROM attachments WHERE approval_type = 'partseries'"
        self.cursor.execute(query_part_series)
        result_part_series = self.cursor.fetchall()
        comp_id_q3 = self.cursor.column_names.index("component_id")
        attach_filename_q3 = self.cursor.column_names.index("attachment_filename")
        
        attachment_dict = {}
        for row1 in self.result_by_part:
            component_ids = []
            for row in result_part_series:
                if row[attach_filename_q3] == row1[self.result_attachmentname]:
                    component_ids.append(row1[self.result_component_id])
                    component_ids.append(row[comp_id_q3])           
            attachment_dict[row1[self.result_attachmentname]] =component_ids
        #print(attachment_dict)

        attachment_dict_filtered = {attachment: component_ids 
                                    for attachment, component_ids in attachment_dict.items() 
                                    if set(component_ids) and len(component_ids) > 0  }
        #print(attachment_dict_filtered)
        
        if attachment_dict_filtered:
            #print("Attachments with Component IDs:")
            logging.info("Attachments with Component IDs:")
            for attachment, component_ids in attachment_dict_filtered.items():
                if(len(component_ids)>0):
                    #print(f"Attachment: {attachment}")
                    logging.info(f"Attachment: {attachment}")
                    #print("Component IDs:")
                    logging.info(("Component IDs:"))
                    for component_id in set(component_ids):
                        #print(component_id)
                        logging.info(component_id)
            count_of_parttseries_bypart=len(attachment_dict_filtered)
            # create a excel with the column names as the attachments and the component ID'S
            df_attachment_dict = pd.DataFrame(attachment_dict_filtered.items(), columns=['Attachment', 'Compnoent IDs'])

            file_partseries=self.path_for_excel/'ByPart_PartSeries_Discrepancy.xlsx'
            file_total=self.path_for_excel/'final_report.xlsx'

            with pd.ExcelWriter(file_partseries) as writer:
                df_attachment_dict.to_excel(writer, index=False)
            
            wb = openpyxl.load_workbook(file_total)
            ws = wb.active
            ws.append(['count of partseries and bypart ',count_of_parttseries_bypart])
            wb.save(file_total)
        
        else:
            #print("No attachments with Component IDs greater than 0.")
            logging.info("No attachments with Component IDs greater than 0.")
        self.count_of_total_links=self.count_of_total_links+count_of_parttseries_bypart

    # find out the link that are both bypart and blanket
    def bypartandblanket(self):
        #print("bypart and blanket")
        logging.info("bypart and blanket")
        count_of_blanketandbypart=0
        # Query to select attachments with approval type 'partseries'
        query_part_series = "SELECT * FROM attachments WHERE approval_type = 'blanket'"
        self.cursor.execute(query_part_series)
        result_bypart_blanket = self.cursor.fetchall()
        comp_id_q3 = self.cursor.column_names.index("component_id")
        attach_filename_q3 = self.cursor.column_names.index("attachment_filename")
        attachment_dict = {}
        for row1 in self.result_by_part:
            component_ids = []
            for row in result_bypart_blanket:
                if row[attach_filename_q3] == row1[self.result_attachmentname]:
                    component_ids.append(row[comp_id_q3])   
                    component_ids.append(row1[self.result_component_id])        
            attachment_dict[row1[self.result_attachmentname]] = component_ids
    
        # Filter attachments with component IDs greater than 0
        attachment_dict_filtered = {attachment: component_ids 
                                    for attachment, component_ids in attachment_dict.items() 
                                    if set(component_ids) and len(component_ids) > 0  }
        #print(attachment_dict_filtered)
        logging.info(attachment_dict_filtered)
        if attachment_dict_filtered:
            #print("Attachments with Component IDs:")
            logging.info("Attachments with Component IDs:")
            for attachment, component_ids in attachment_dict_filtered.items():
                # print(f"Attachment: {attachment}")
                logging.info(f"Attachment: {attachment}")
                # print("Component IDs:")
                logging.info("component IDs:")
                for component_id in set(component_ids):
                    # print(component_id)
                    logging.info(component_id)
            for attachment,component_ids in attachment_dict_filtered.items():
                attachment_dict_filtered[attachment]=set(component_ids)
                    
            count_of_blanketandbypart=len(attachment_dict_filtered)
            df_attachment_dict = pd.DataFrame(attachment_dict_filtered.items(), columns=['Attachment', 'Compnoent IDs'])
            # Save DataFrame to Excel file 'duplicate_byparts.xlsx'
            file_blanket=self.path_for_excel/'ByPart_Blanket_Discrepancy.xlsx'
            file_total=self.path_for_excel/'final_report.xlsx'


            with pd.ExcelWriter(file_blanket) as writer:
                df_attachment_dict.to_excel(writer, index=False)
                
            wb = openpyxl.load_workbook(file_total)
            ws = wb.active
            ws.append(['count of partseries and bypart ',count_of_blanketandbypart])
            wb.save(file_total)
            
           # print("done with bypart blanket")
            logging.info("done with bypart blanket")
        else:
            print("No attachments with Component IDs greater than 0.")
        self.count_of_total_links=self.count_of_total_links+count_of_blanketandbypart
            
            
    def getlinks(self):
       # print(self.attachment_dict)
        for attachment,count in self.attachment_dict.items():
            self.list_of_duplicate_byparts.append(attachment)
       # print(self.list_of_duplicate_byparts)    
    #clear the excel sheet before opening to prevent it from overriding
    def clear_excel_file(self,file_path):
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None
        wb.save(file_path)


if __name__ == "__main__":
    print("ENTER THE VALID INPUT:")
    print("**********************")
    print(" MENU :")
    print("1: validate")
    print("2: Validate and download")
    print("***********************")
    choice = int(input("ENTER YOUR CHOICE :"))
    try:
        host = input("Enter the host details of the Database: ")
        user = input("Enter the user details of the Database: ")
        password = input("Enter the  password of the Database: ")
        database = input("Enter the name of the database: ")
        db_connection = DatabaseConnection(host=host, user=user, password=password, database=database)
        # db_connection = DatabaseConnection(host="localhost", user="root", password="", database="icm_db")
        db_connection.connect()
        obj = Connect(db_connection.conn)

    # print("entered")

        def choose(choice):
            if choice == 1:
                validate()
            elif choice == 2:
                validate()
                download()
            else:
                print("Invalid choice")

        def validate():
            obj.process_attachments()
            obj.duplicate_byparts()
            obj.getlinks()
            obj.count_of_linktypes()
            obj.partseries()
            obj.bypartandblanket()

        def download():
            obj.download_all_file()

        choose(choice)
        db_connection.close()
    except Exception as e:
        print("error is:",e)

