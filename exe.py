#python version 3.12.2
#requests Version: 2.31.0
#mysql.connector  version 8.3.0
#pandas version 2.2.1
#openpyxl version 3.1.2
#pathlib version 1.0.1


import mysql.connector                     #this module is for connection esablishment
import requests                            # this is to fetch the requests
from collections import Counter            # to count the no of time s a particular link occurs
from pathlib import Path                   # to check the path
import pandas as pd                        # to handel dataframes to import into excel files
import openpyxl                            # to handel and creation of excel files

# Connection establishment
try:
    conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="icm_db"
    )
    print("Connection established successfully!")
except mysql.connector.Error as err:
    print(f"Error: {err}")

cursor = conn.cursor()

# Selecting data from the attachments table
query = "SELECT * FROM attachments"
cursor.execute(query)
# Fetching all records in the database
result = cursor.fetchall()
# Lists to store links and paths
m_list = []
link = []
path = []
attachment_dict = {}
list_of_duplicate_byparts=[]

# Getting the column indices
column_index = cursor.column_names.index('attachment_filename')
attachments_type = cursor.column_names.index('approval_type')
comp_id_attachments=cursor.column_names.index("component_id")

# Loop to check whether it's a link or a path
count_of_total_links=0
for row in result:
    count_of_total_links=count_of_total_links+1
    if row[column_index].startswith("https://") or row[column_index].startswith("http://"):
        link.append(row[column_index])
    else:
        path.append(row[column_index])

# Function to download files
def download_file(file_url, local_directory):
    try:
        # Get request
        response = requests.get(file_url)
        #split the query parameter if exists else it takes the link only and append that to the path parameter
        filename = local_directory / Path(file_url.split('?')[0]).name.strip()
        
        # If the file exists
        if filename.exists():
            print("File already exists:")
            return True

        # If not present, create it
        with open(filename, 'wb') as local_file:
            local_file.write(response.content)
        
        print(f"File downloaded and saved as {filename}")
        parts=file_url.parts
        index=parts.index("atom_mfr_docs")
        list_of_parts=list(index[4:])
        final_path=str(list_of_parts)
        final_path=Path(final_path)
        
        
        qd="update attachments set attachment_filename=%s"
        
        cursor.execute(qd,(final_path))
        return True
    except requests.exceptions.RequestException as e:
        print(f"Error downloading file: {e}")
        return False
    
# Query to execute the manufacturers table
query1 = "SELECT * FROM manufacturers"
cursor.execute(query1)
manufacture_results = cursor.fetchall()
#column_names.index is a method to access the index of the particular colum as the list item cant be accessed by its 
manufacture_name_manufacture = cursor.column_names.index('name')
manufucture_index = cursor.column_names.index('normalized_name')
manufacturer_id_index = cursor.column_names.index('id')

# Query to execute the master_components table
query2 = "SELECT * FROM master_components"
cursor.execute(query2)
component_results = cursor.fetchall()
manufacture_part_in_master =cursor.column_names.index('manufacturer_partnumber')
component_id_component = cursor.column_names.index('id')
manufacture_name_component =cursor.column_names.index('manufacturer_name')

def download_all_file():
    global m_list
    print("entered download_all_files")
    print(list_of_duplicate_byparts)
    print(m_list)
    pat = input("Enter the path you want to insert the folders:") 
    try:
        for row in list_of_duplicate_byparts:
            q="select *  from attachments where attachment_filename=%s  group by manufacturer_id "
            cursor.execute(q,(row,))
            res=cursor.fetchall()
            if len(res)>1:
                for i in res:
                    dir_name_multiple = "compilance"
                    local_directory_multiple = Path(pat) / dir_name_multiple
                    if local_directory_multiple.exists():
                        print("already present multiple fiolder")
                    else:
                        local_directory_multiple.mkdir()
                        print("Directory for multiple files was created :", local_directory_multiple)
                    download_file(row, local_directory_multiple)
                    m_list.append(row)
                
        
        for row in manufacture_results:
            manufacturer_id = row[manufacturer_id_index]
            directory_name = row[manufucture_index]
            manu_name = row[manufacture_name_manufacture]
            local_directory = Path(pat) / directory_name
            
            if local_directory.exists():
                print("already exists:", local_directory)
            # If not present, create it
            else:
                local_directory.mkdir()
                print("Directory created:", local_directory)
                
            # Compilance directory
            directory_name_comp = 'compilance'
            local_directory_compilance = local_directory / directory_name_comp
            
            # if the directory exists, enter it
            if local_directory_compilance.exists():
                print("already present compilance directory")
            # else create it
            else:
                local_directory_compilance.mkdir()
                print("Compilance directory created:", local_directory_compilance)
            
            query2 = "SELECT * FROM master_components WHERE manufacturer_name = %s"
            cursor.execute(query2, (manu_name,))
            component_results = cursor.fetchall()
            
            manufacture_part_in_master = cursor.column_names.index('manufacturer_partnumber')
            component_id_component = cursor.column_names.index('id')
            manufacture_name_component = cursor.column_names.index('manufacturer_name')
        
            for component_row in component_results:
                if component_row[manufacture_name_component] == row[manufacture_name_manufacture]:
                    component_id = component_row[component_id_component] 
                    directory_name1 = component_row[manufacture_part_in_master] 
                    local_directory2 = local_directory_compilance / directory_name1

                    if local_directory2.exists():
                        print("Entered component part folder already exists")
                    else:    
                        local_directory2.mkdir(parents=True, exist_ok=True)
                        print(f"Subdirectory created: {local_directory2}")
                    #print(list_of_duplicate_byparts)
                    if list_of_duplicate_byparts:
                        attachment_placeholders = ', '.join(['%s'] * len(list_of_duplicate_byparts))
                        query3 = f"SELECT * FROM attachments WHERE manufacturer_id = %s AND component_id = %s AND attachment_filename NOT IN ({attachment_placeholders})"
                        query_params = [manufacturer_id, component_id] + list_of_duplicate_byparts
                    #    print("query_params are")
                    #    print(query_params)
                        cursor.execute(query3, query_params)
                    #    print("query is:")
                    #    print(query3)
                        attachment_results = cursor.fetchall()
                    #    print(attachment_results)
                        
                        att_res_lin = cursor.column_names.index("attachment_filename")
                    
                        for attachment_row in attachment_results:
                            file_url = attachment_row[att_res_lin]
                            file_type = attachment_row[attachments_type]

                            if file_type == "partseries":
                                directory_name_partseries = 'partseries'
                                local_directory_partseries = local_directory_compilance / directory_name_partseries
                                if not local_directory_partseries.exists():
                                    local_directory_partseries.mkdir()
                                    print("Partseries directory created:", local_directory_partseries)
                                download_file(file_url, local_directory_partseries)
                            elif file_type == "blanket":
                                directory_name_blanket = 'blanket'
                                local_directory_blanket = local_directory_compilance / directory_name_blanket

                                if not local_directory_blanket.exists():
                                    local_directory_blanket.mkdir()
                                    print("Blanket directory created:", local_directory_blanket)

                                download_file(file_url, local_directory_blanket)
                            else:
                                download_file(file_url, local_directory2)
        m_list=set(m_list)
        print(m_list)
        for k in m_list:
            list_of_duplicate_byparts.remove(k)
        print(list_of_duplicate_byparts)
        for i in list_of_duplicate_byparts:
            print(i)
            qp="select * from attachments where approval_type='By part' and attachment_filename=%s group by manufacturer_id"
            cursor.execute(qp,(i,))
            m_id=cursor.column_names.index("manufacturer_id")
            ans=cursor.fetchall()

            for j in ans:
                aq="select * from manufacturers where id=%s"
                cursor.execute(aq,(m_id,))
                ans_in_j=cursor.fetchall()
                n=cursor.column_names.index("normalized_name")
                for k in ans_in_j:
                    n_name=k[n]
                    com="compilance"
                    nam="multiplebyparts"
                    dir=Path(pat) / n_name/ com / nam
                    if dir.exists():
                        print("already multiple exists")
                    else:
                        dir.mkdir()
                        print("created")
                    download_file(i,dir)
                                    
                                
                                
    except Exception as e:
         print(f"Error: {e}")
    # # finally:
    # #     for r in m:
    # #         list_of_duplicate_byparts.remove(r)
        
    # #     print("final part")
    # #     print(list_of_duplicate_byparts)







        
def count_of_linktypes():
    dict_count_types={}
    dict_count_links={}

    for i in range(len(link)):
        count_bypart = 0
        count_partseries = 0
        
        for row in result:
            if row[column_index] == link[i]:

                if row[attachments_type] == 'By part':
                    count_bypart += 1
                elif row[attachments_type] == 'partseries':
                    count_partseries += 1
        dict_count_types[link[i]]={'By part': count_bypart,'Partseries': count_partseries}
   
    #counter is a function which counts the occurance of the each item 
    link_count = Counter(link)
    for j, count in link_count.items():
        dict_count_links[j]=count
        
    df_link_count = pd.DataFrame(dict_count_links.items(), columns=['Link', 'Count'])
    #to convert the dictonary data into pandas dataframepytho
    df_link_types = pd.DataFrame.from_dict(dict_count_types, orient='index').reset_index()
    
    
    # Save data to Excel
    with pd.ExcelWriter('count_of_linktypes.xlsx') as writer:
        df_link_count.to_excel(writer, sheet_name='Link Counts', index=False)
        df_link_types.to_excel(writer, sheet_name='Link Details', index=False)
            

def duplicate_byparts():
    print("duplicate bypart")
    count_of_duplicatebyparts = 0
    
    
    # Query to fetch all attachments
    query_on_duplicate = "SELECT * FROM attachments"
    cursor.execute(query_on_duplicate)
    results_of_duplicates = cursor.fetchall()
    
    # Indxes for those compo
    results_of_duplicates_attachment_file = cursor.column_names.index("attachment_filename")
    results_of_duplicates_approval = cursor.column_names.index("approval_type")
    results_of_duplicates_id = cursor.column_names.index("component_id")
    
    for link_item in link:
        for row in results_of_duplicates:
            if row[results_of_duplicates_attachment_file] == link_item and row[results_of_duplicates_approval] == "By part":
                attachment_filename = row[results_of_duplicates_attachment_file]
                
                # Query to fetch component IDs for the attachment
                query_component_id = "SELECT component_id FROM attachments WHERE attachment_filename = %s AND approval_type = 'By part'"
                cursor.execute(query_component_id, (attachment_filename,))
                result_component_ids = cursor.fetchall()
                
                # Extract component IDs
                component_ids = [res[0] for res in result_component_ids]

                # Check if there are duplicate component IDs
                if len(set(component_ids)) > 1:
                    attachment_dict[attachment_filename] = component_ids
    for attachment,component_id in attachment_dict.items():
        attachment_dict[attachment]=set(component_id)
    print(attachment_dict)
    count_of_duplicatebyparts=len(attachment_dict)
    if attachment_dict:
        # Create DataFrame from attachment dictionary
        df_attachment_dict = pd.DataFrame(attachment_dict.items(), columns=['Attachment', 'Component IDs'])
        df_duplicate_bypart = pd.DataFrame([['duplicate by parts', count_of_duplicatebyparts]], columns=['Attachment', 'Count'])
        df_total_links=pd.DataFrame([['total links',count_of_total_links]],columns=['Attachment', 'Count'])
        count_of_uniquelinks=len(set(link))
        df_uniquelinks=pd.DataFrame([['unique links',count_of_uniquelinks]],columns=['Attachment', 'Count'])
        
        # Save DataFrame to Excel file 'duplicate_byparts.xlsx'
        with pd.ExcelWriter('duplicate_byparts.xlsx') as writer:
            df_attachment_dict.to_excel(writer, index=False)
        filename=Path(r'C:\Users\mohanvamsi.sadhu\Desktop\dem\total.xlsx')
        if filename.exists():
            clear_excel_file(filename)
            wb = openpyxl.load_workbook('total.xlsx')
            ws = wb.active
            
            ws.append(['count of total links', len(link)]) 
            ws.append(['count of unique links',count_of_uniquelinks])
            ws.append(['count of duplicate by parts', count_of_duplicatebyparts])
            wb.save('total.xlsx')
            
        else:
            with pd.ExcelWriter('total.xlsx') as writer:
                df_total_links.to_excel(writer, index=False)
            wb = openpyxl.load_workbook('total.xlsx')
            ws = wb.active
            ws.append(['count of unique links ',count_of_uniquelinks])
            ws.append(['count of duplicate by parts',count_of_duplicatebyparts])
            wb.save('total.xlsx')
               

#function for the same link has more than than one approval types as bypart and the part series
def bypartandblanket():
    count_of_blanketandbypart=0
    print("bypart and blanket")
    count_of_blanketandbypart=0
    query_by_part = "SELECT * FROM attachments WHERE approval_type = 'By part'"
    cursor.execute(query_by_part)
    result_by_part = cursor.fetchall()
    result_id = cursor.column_names.index("component_id")
    result_attachmentname = cursor.column_names.index("attachment_filename")
    # Query to select attachments with approval type 'partseries'
    query_part_series = "SELECT * FROM attachments WHERE approval_type = 'blanket'"
    cursor.execute(query_part_series)
    result_part_series = cursor.fetchall()
    comp_id_q3 = cursor.column_names.index("component_id")
    attach_filename_q3 = cursor.column_names.index("attachment_filename")
    attachment_dict = {}
    for row1 in result_by_part:
        component_ids = []
        for row in result_part_series:
            if row[attach_filename_q3] == row1[result_attachmentname]:
                component_ids.append(row[comp_id_q3])   
                component_ids.append(row1[result_id])        
        attachment_dict[row1[result_attachmentname]] = component_ids
  
    # Filter attachments with component IDs greater than 0
    attachment_dict_filtered = {attachment: component_ids 
                                for attachment, component_ids in attachment_dict.items() 
                                if set(component_ids) and len(component_ids) > 0  }
    print(attachment_dict_filtered)
    if attachment_dict_filtered:
        print("Attachments with Component IDs:")
        for attachment, component_ids in attachment_dict_filtered.items():
            print(f"Attachment: {attachment}")
            print("Component IDs:")
            for component_id in set(component_ids):
                print(component_id)
        for attachment,component_ids in attachment_dict_filtered.items():
            attachment_dict_filtered[attachment]=set(component_ids)
                
        count_of_blanketandbypart=len(attachment_dict_filtered)
        df_attachment_dict = pd.DataFrame(attachment_dict_filtered.items(), columns=['Attachment', 'Compnoent IDs'])
        df_counts=pd.DataFrame([['blanket and partseries', count_of_blanketandbypart]], columns=['Attachment', 'Count'])
        # Save DataFrame to Excel file 'duplicate_byparts.xlsx'
        with pd.ExcelWriter('bypartblanket.xlsx') as writer:
            df_attachment_dict.to_excel(writer, index=False)
            
        wb = openpyxl.load_workbook('total.xlsx')
        ws = wb.active
        ws.append(['count of partseries and bypart ',count_of_blanketandbypart])
        wb.save('total.xlsx')
        
        print("done with bypart blanket")
    else:
        print("No attachments with Component IDs greater than 0.")

#function for the more than 1 link has both the by part and the partseries
def partseriesbypart():
    count_of_parttseries_bypart=0
    print("partseries")
    query_by_part = "SELECT * FROM attachments WHERE approval_type = 'By part'"
    cursor.execute(query_by_part)
    result_by_part = cursor.fetchall()
    result_attachmentname = cursor.column_names.index("attachment_filename")
    result_component_id=cursor.column_names.index("component_id")
    
    # Query to select attachments with approval type 'partseries'
    query_part_series = "SELECT * FROM attachments WHERE approval_type = 'partseries'"
    cursor.execute(query_part_series)
    result_part_series = cursor.fetchall()
    comp_id_q3 = cursor.column_names.index("component_id")
    attach_filename_q3 = cursor.column_names.index("attachment_filename")
    
    attachment_dict = {}
    for row1 in result_by_part:
        component_ids = []
        for row in result_part_series:
            if row[attach_filename_q3] == row1[result_attachmentname]:
                component_ids.append(row1[result_component_id])
                component_ids.append(row[comp_id_q3])           
        attachment_dict[row1[result_attachmentname]] =component_ids
    print(attachment_dict)

    attachment_dict_filtered = {attachment: component_ids 
                                for attachment, component_ids in attachment_dict.items() 
                                if set(component_ids) and len(component_ids) > 0  }
    print(attachment_dict_filtered)
    
    if attachment_dict_filtered:
        print("Attachments with Component IDs:")
        for attachment, component_ids in attachment_dict_filtered.items():
            if(len(component_ids)>0):
                print(f"Attachment: {attachment}")
                print("Component IDs:")
                for component_id in set(component_ids):
                    print(component_id)
        count_of_parttseries_bypart=len(attachment_dict_filtered)
        # create a excel with the column names as the attachments and the component ID'S
        df_attachment_dict = pd.DataFrame(attachment_dict_filtered.items(), columns=['Attachment', 'Compnoent IDs'])
        with pd.ExcelWriter('partseriesbypart.xlsx') as writer:
            df_attachment_dict.to_excel(writer, index=False)
        
        wb = openpyxl.load_workbook('total.xlsx')
        ws = wb.active
        ws.append(['count of partseries and bypart ',count_of_parttseries_bypart])
        wb.save('total.xlsx')
    
    else:
        print("No attachments with Component IDs greater than 0.")



def getlinks():
    print(attachment_dict)
    for attachment,count in attachment_dict.items():
        list_of_duplicate_byparts.append(attachment)
    print(list_of_duplicate_byparts)

def clear_excel_file(file_path):
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None
    wb.save(file_path)

def validate():
    duplicate_byparts()
    getlinks()
    count_of_linktypes()
    partseriesbypart()
    bypartandblanket()
        
def selection(key):
    if key==1:
        validate()
    elif key==2:
        validate()
        download_all_file()
    else:
        print("invalid input")
        
print("ENTER THE VALID INPUT:")
print("**********************")
print(" MENU :")
print("1: validate")
print("2: Validate and download")
print("***********************")
inp=int(input("ENTER YOUR CHOICE :"))
selection(inp)





