import mysql.connector
import requests
from collections import Counter
from pathlib import Path
import pandas as pd
import openpyxl

class DatabaseConnection:
    def __init__(self, host, user, password, database):
        self.host = host
        self.user = user
        self.password = password
        self.database = database
        self.conn = None

    def connect(self):
        try:
            self.conn = mysql.connector.connect(
                host=self.host,
                user=self.user,
                password=self.password,
                database=self.database
            )
            print("Connection established successfully!")
        except mysql.connector.Error as err:
            print(f"Error: {err}")

    def close(self):
        if self.conn:
            self.conn.close()

class Connect:
    def __init__(self, conn):
        self.conn = conn
        self.cursor = self.conn.cursor()
        self.link = []
        self.path = []
        self.attachment_dict = {}
        self.list_of_duplicate_byparts = []
        self.m_list=[]

    def process_attachments(self):
        query = "SELECT * FROM attachments"
        self.cursor.execute(query)
        self.result = self.cursor.fetchall()
        self.column_index = self.cursor.column_names.index('attachment_filename')
        self.attachments_type = self.cursor.column_names.index('approval_type')
        
        
        query_by_part = "SELECT * FROM attachments WHERE approval_type = 'By part'"
        self.cursor.execute(query_by_part)
        self.result_by_part = self.cursor.fetchall()
        self.result_attachmentname = self.cursor.column_names.index("attachment_filename")
        self.result_component_id=self.cursor.column_names.index("component_id")
        
        
        query1 = "SELECT * FROM manufacturers"
        self.cursor.execute(query1)
        self.manufacture_results = self.cursor.fetchall()
        #column_names.index is a method to access the index of the particular colum as the list item cant be accessed by its 
        self.manufacture_name_manufacture = self.cursor.column_names.index('name')
        self.manufucture_index = self.cursor.column_names.index('normalized_name')
        self.manufacturer_id_index = self.cursor.column_names.index('id')

        # Query to execute the master_components table
        query2 = "SELECT * FROM master_components"
        self.cursor.execute(query2)
        self.component_results = self.cursor.fetchall()
        self.manufacture_part_in_master =self.cursor.column_names.index('manufacturer_partnumber')
        self.component_id_component = self.cursor.column_names.index('id')
        self.manufacture_name_component =self.cursor.column_names.index('manufacturer_name')

        for row in self.result:
            if row[self.column_index].startswith("https://") or row[self.column_index].startswith("http://"):
                self.link.append(row[self.column_index])
            else:
                self.path.append(row[self.column_index])

        self.count_of_total_links = len(self.link)

    def download_all_file(self):
        global m_list
        print("entered download_all_files")
        # print(self.list_of_duplicate_byparts)
        # print(self.m_list)
        global pat 
        pat = input("Enter the path you want to insert the folders:") 
        try:
            print(self.list_of_duplicate_byparts)
            for row in self.list_of_duplicate_byparts:
                q="select *  from attachments where attachment_filename=%s and approval_type='By part' group by manufacturer_id "
                self.cursor.execute(q,(row,))
                res=self.cursor.fetchall()
                a_id=self.cursor.column_names.index("id")
                afn=self.cursor.column_names.index("attachment_filename")
                a_mid=self.cursor.column_names.index("manufacturer_id")
                print(res)
                print("dem")
                if len(res)>1:
                    for m in res:
                        id_for_attachment=m[a_id]
                        dem=m[afn]
                        attachment_mid=m[a_mid]
                        que="select * from attachments where attachment_filename=%s and manufacturer_id=%s and approval_type='By part'"
                        self.cursor.execute(que,(dem,attachment_mid))
                        res_attachment=self.cursor.fetchall()
                        for i in res_attachment:
                            dir_name_multiple = "compilance"
                            local_directory_multiple = Path(pat) / dir_name_multiple
                            if local_directory_multiple.exists():
                                print("already present multiple fiolder")
                            else:
                                local_directory_multiple.mkdir()
                                print("Directory for multiple files was created :", local_directory_multiple)
                            self.download_file(dem, local_directory_multiple,id_for_attachment)
                            self.m_list.append(dem)
                        
            
            for row in self.manufacture_results:
                print("y")
                manufacturer_id = row[self.manufacturer_id_index]
                directory_name = row[self.manufucture_index]
                manu_name = row[self.manufacture_name_manufacture]
                local_directory = Path(pat) / directory_name
                
                if local_directory.exists():
                    print("already exists:", local_directory)
                # If not present, create it
                else:
                    local_directory.mkdir()
                    print("Directory created:", local_directory)
                    
                # Compilance directory
                directory_name_comp = 'compilance'
                local_directory_compilance = local_directory / directory_name_comp
                
                # if the directory exists, enter it
                if local_directory_compilance.exists():
                    print("already present compilance directory")
                # else create it
                else:
                    local_directory_compilance.mkdir()
                    print("Compilance directory created:", local_directory_compilance)
                
                query2 = "SELECT * FROM master_components WHERE manufacturer_name = %s"
                self.cursor.execute(query2, (manu_name,))
                component_results = self.cursor.fetchall()
                
                manufacture_part_in_master = self.cursor.column_names.index('manufacturer_partnumber')
                component_id_component = self.cursor.column_names.index('id')
                manufacture_name_component = self.cursor.column_names.index('manufacturer_name')
            
                for component_row in component_results:
                    if component_row[manufacture_name_component] == row[self.manufacture_name_manufacture]:
                        component_id = component_row[component_id_component] 
                        directory_name1 = component_row[manufacture_part_in_master] 
                        local_directory2 = local_directory_compilance / directory_name1

                        if local_directory2.exists():
                            print("Entered component part folder already exists")
                        else:    
                            local_directory2.mkdir(parents=True, exist_ok=True)
                            print(f"Subdirectory created: {local_directory2}")
                        #print(self.list_of_duplicate_byparts)
                        if self.list_of_duplicate_byparts:
                            attachment_placeholders = ', '.join(['%s'] * len(self.list_of_duplicate_byparts))
                            query3 = f"SELECT * FROM attachments WHERE manufacturer_id = %s AND component_id = %s AND attachment_filename NOT IN ({attachment_placeholders})"
                            query_params = [manufacturer_id, component_id] + self.list_of_duplicate_byparts
                        #    print("query_params are")
                        #    print(query_params)
                            self.cursor.execute(query3, query_params)
                        #    print("query is:")
                        #    print(query3)
                            attachment_results = self.cursor.fetchall()
                            print("Column names:", self.cursor.column_names)
                            id_of_index=self.cursor.column_names.index('id')
                        #    print(attachment_results)
                            
                            att_res_lin = self.cursor.column_names.index("attachment_filename")
                        
                            for attachment_row in attachment_results:
                                file_url = attachment_row[att_res_lin]
                                file_type = attachment_row[self.attachments_type]
                                id_of_a= attachment_row[id_of_index]

                                if file_type == "partseries":
                                    directory_name_partseries = 'partseries'
                                    local_directory_partseries = local_directory_compilance / directory_name_partseries
                                    if not local_directory_partseries.exists():
                                        local_directory_partseries.mkdir()
                                        print("Partseries directory created:", local_directory_partseries)
                                        print("is_of_a",id_of_a)
                                        #print("download")
                                    self.download_file(file_url, local_directory_partseries,id_of_a)
                                elif file_type == "blanket":
                                    directory_name_blanket = 'blanket'
                                    local_directory_blanket = local_directory_compilance / directory_name_blanket

                                    if not local_directory_blanket.exists():
                                        local_directory_blanket.mkdir()
                                        print("Blanket directory created:", local_directory_blanket)
                                    #print(download)
                                    print("id_of_a",id_of_a)
                                    self.download_file(file_url, local_directory_blanket,id_of_a)
                                else:
                                    #print("download")
                                    print(id_of_a)
                                    self.download_file(file_url, local_directory2,id_of_a)
            self.m_list=set(self.m_list)
            print(self.m_list)
            for k in self.m_list:
                self.list_of_duplicate_byparts.remove(k)
            print(self.list_of_duplicate_byparts)
            for i in self.list_of_duplicate_byparts:
                print(i)
                qp="select * from attachments where approval_type='By part' and attachment_filename=%s"
                self.cursor.execute(qp,(i,))
                m_id=self.cursor.column_names.index("manufacturer_id")
                ans=self.cursor.fetchall()
                id_in_index=self.cursor.column_names.index("id")

                for j in ans:
                    id_in_attachments=j[id_in_index]
                    aq="select * from manufacturers where id=%s"
                    self.cursor.execute(aq,(j[m_id],))
                    ans_in_j=self.cursor.fetchall()
                    n=self.cursor.column_names.index("normalized_name")
                    for k in ans_in_j:
                        n_name=k[n]
                        com="compilance"
                        nam="multiplebyparts"
                        dir=Path(pat) / n_name/ com / nam
                        if dir.exists():
                            print("already multiple exists")
                        else:
                            dir.mkdir()
                            print("created")
                        print("id_in_attachments",id_in_attachments)
                        self.download_file(i,dir,id_in_attachments)                  
                                    
                                    
        except Exception as e:
            print(f"Error: {e}")
            
    def download_file(self, file_url, local_directory, id_of_attachments):
        try:
            # Download the file from the URL
            response = requests.get(file_url)
            filename = local_directory / Path(file_url.split('?')[0]).name.strip()

            # Check if file already exists
            if filename.exists():
                print("File already exists:", filename)
                pa = Path(local_directory) / filename
                part = pa.parts
                index = part.index("atom_mfr_docs")
                list_of_parts = list(part[index:])
                finalpath = "\\".join(list_of_parts)
                print(id_of_attachments)
                
                # Update the database record with the path of the downloaded file
                query_to_update = "UPDATE attachments SET notes=%s WHERE id=%s"
                self.cursor.execute(query_to_update, (finalpath, id_of_attachments))
                print("File path updated in the database:", finalpath)
                self.conn.commit()
                return True

            # Save the file to the local directory
            with open(filename, 'wb') as local_file:
                local_file.write(response.content)
            print(f"File downloaded and saved as {filename}")

            # Manipulate file path to get the relative path from 'atom_mfr_docs'
            pa = Path(local_directory) / filename
            part = pa.parts
            index = part.index("atom_mfr_docs")
            list_of_parts = list(part[index:])
            finalpath = "\\".join(list_of_parts)
            print(id_of_attachments)

            # Update the database record with the path of the downloaded file
            query_to_update = "UPDATE attachments SET notes=%s WHERE id=%s"
            self.cursor.execute(query_to_update, (finalpath, id_of_attachments))
            print("File path updated in the database:", finalpath)

            # Commit the transaction
            self.conn.commit()
            
            return True  # Return True to indicate successful download and database update

        except Exception as e:
            print("An exception occurred:", e)
            return False  # Return False to indicate failure
        
                        
    def count_of_linktypes(self):
        dict_count_links = {}
        for link_item in self.link:
            count_bypart = 0
            count_partseries = 0
            for row in self.result:
                if row[self.column_index] == link_item:
                    if row[self.attachments_type] == 'By part':
                        count_bypart += 1
                    elif row[self.attachments_type] == 'partseries':
                        count_partseries += 1
            dict_count_links[link_item] = {'By part': count_bypart, 'Partseries': count_partseries}

        link_count = Counter(self.link)
        for j, count in link_count.items():
            dict_count_links[j] = count

        df_link_count = pd.DataFrame(dict_count_links.items(), columns=['Link', 'Count'])
        df_link_types = pd.DataFrame.from_dict(dict_count_links, orient='index').reset_index()

        with pd.ExcelWriter('count_of_linktypes.xlsx') as writer:
            df_link_count.to_excel(writer, sheet_name='Link Counts', index=False)
            df_link_types.to_excel(writer, sheet_name='Link Details', index=False)

    def duplicate_byparts(self):
        self.attachment_dict = {}
        self.count_of_duplicatebyparts=0
        for link in self.link:
            for row in self.result:
                if row[self.attachments_type] == 'By part' and row[self.column_index]==link:
                    attachment_filename=row[self.column_index]
                    query_component_id = "SELECT component_id FROM attachments WHERE attachment_filename = %s AND approval_type = 'By part'"
                    self.cursor.execute(query_component_id, (attachment_filename,))
                    result_component_ids = self.cursor.fetchall()
                
                     # Extract component IDs
                    component_ids = [res[0] for res in result_component_ids]

                # Check if there are duplicate component IDs
                    if len(set(component_ids)) > 1:
                        self.attachment_dict[attachment_filename] = component_ids
        for attachment,component_id in self.attachment_dict.items():
            self.attachment_dict[attachment]=set(component_id)
        self.count_of_duplicatebyparts=len(self.attachment_dict)

        if self.attachment_dict:
            df_attachment_dict = pd.DataFrame(self.attachment_dict.items(), columns=['Attachment', 'Component IDs'])
            df_duplicate_bypart = pd.DataFrame([['duplicate by parts', self.count_of_duplicatebyparts]],
                                               columns=['Attachment', 'Count'])
            df_total_links = pd.DataFrame([['total links', self.count_of_total_links]],
                                          columns=['Attachment', 'Count'])
            count_of_uniquelinks = len(set(self.link))
            df_uniquelinks = pd.DataFrame([['unique links', count_of_uniquelinks]],
                                          columns=['Attachment', 'Count'])

            with pd.ExcelWriter('duplicate_byparts.xlsx') as writer:
                df_attachment_dict.to_excel(writer, index=False)

            filename = Path(r'C:\Users\mohanvamsi.sadhu\Desktop\dem\total.xlsx')
            if filename.exists():
                self.clear_excel_file(filename)
                wb = openpyxl.load_workbook('total.xlsx')
                ws = wb.active

                ws.append(['count of total links', len(self.link)])
                ws.append(['count of unique links', count_of_uniquelinks])
                ws.append(['count of duplicate by parts', self.count_of_duplicatebyparts])
                wb.save('total.xlsx')

            else:
                with pd.ExcelWriter('total.xlsx') as writer:
                    df_total_links.to_excel(writer, index=False)
                wb = openpyxl.load_workbook('total.xlsx')
                ws = wb.active
                ws.append(['count of unique links ', count_of_uniquelinks])
                ws.append(['count of duplicate by parts', self.count_of_duplicatebyparts])
                wb.save('total.xlsx')
                
                
    def partseries(self):
        self.count_of_partseries_bypart=0
        print("partseries")

        
        # Query to select attachments with approval type 'partseries'
        query_part_series = "SELECT * FROM attachments WHERE approval_type = 'partseries'"
        self.cursor.execute(query_part_series)
        result_part_series = self.cursor.fetchall()
        comp_id_q3 = self.cursor.column_names.index("component_id")
        attach_filename_q3 = self.cursor.column_names.index("attachment_filename")
        
        attachment_dict = {}
        for row1 in self.result_by_part:
            component_ids = []
            for row in result_part_series:
                if row[attach_filename_q3] == row1[self.result_attachmentname]:
                    component_ids.append(row1[self.result_component_id])
                    component_ids.append(row[comp_id_q3])           
            attachment_dict[row1[self.result_attachmentname]] =component_ids
        #print(attachment_dict)

        attachment_dict_filtered = {attachment: component_ids 
                                    for attachment, component_ids in attachment_dict.items() 
                                    if set(component_ids) and len(component_ids) > 0  }
        #print(attachment_dict_filtered)
        
        if attachment_dict_filtered:
            print("Attachments with Component IDs:")
            for attachment, component_ids in attachment_dict_filtered.items():
                if(len(component_ids)>0):
                    print(f"Attachment: {attachment}")
                    print("Component IDs:")
                    for component_id in set(component_ids):
                        print(component_id)
            count_of_parttseries_bypart=len(attachment_dict_filtered)
            # create a excel with the column names as the attachments and the component ID'S
            df_attachment_dict = pd.DataFrame(attachment_dict_filtered.items(), columns=['Attachment', 'Compnoent IDs'])
            with pd.ExcelWriter('partseriesbypart.xlsx') as writer:
                df_attachment_dict.to_excel(writer, index=False)
            
            wb = openpyxl.load_workbook('total.xlsx')
            ws = wb.active
            ws.append(['count of partseries and bypart ',count_of_parttseries_bypart])
            wb.save('total.xlsx')
        
        else:
            print("No attachments with Component IDs greater than 0.")
    def bypartandblanket(self):
        print("bypart and blanket")
        count_of_blanketandbypart=0
        # Query to select attachments with approval type 'partseries'
        query_part_series = "SELECT * FROM attachments WHERE approval_type = 'blanket'"
        self.cursor.execute(query_part_series)
        result_part_series = self.cursor.fetchall()
        comp_id_q3 = self.cursor.column_names.index("component_id")
        attach_filename_q3 = self.cursor.column_names.index("attachment_filename")
        attachment_dict = {}
        for row1 in self.result_by_part:
            component_ids = []
            for row in result_part_series:
                if row[attach_filename_q3] == row1[self.result_attachmentname]:
                    component_ids.append(row[comp_id_q3])   
                    component_ids.append(row1[self.result_component_id])        
            attachment_dict[row1[self.result_attachmentname]] = component_ids
    
        # Filter attachments with component IDs greater than 0
        attachment_dict_filtered = {attachment: component_ids 
                                    for attachment, component_ids in attachment_dict.items() 
                                    if set(component_ids) and len(component_ids) > 0  }
        print(attachment_dict_filtered)
        if attachment_dict_filtered:
            print("Attachments with Component IDs:")
            for attachment, component_ids in attachment_dict_filtered.items():
                print(f"Attachment: {attachment}")
                print("Component IDs:")
                for component_id in set(component_ids):
                    print(component_id)
            for attachment,component_ids in attachment_dict_filtered.items():
                attachment_dict_filtered[attachment]=set(component_ids)
                    
            count_of_blanketandbypart=len(attachment_dict_filtered)
            df_attachment_dict = pd.DataFrame(attachment_dict_filtered.items(), columns=['Attachment', 'Compnoent IDs'])
            df_counts=pd.DataFrame([['blanket and partseries', count_of_blanketandbypart]], columns=['Attachment', 'Count'])
            # Save DataFrame to Excel file 'duplicate_byparts.xlsx'
            with pd.ExcelWriter('bypartblanket.xlsx') as writer:
                df_attachment_dict.to_excel(writer, index=False)
                
            wb = openpyxl.load_workbook('total.xlsx')
            ws = wb.active
            ws.append(['count of partseries and bypart ',count_of_blanketandbypart])
            wb.save('total.xlsx')
            
            print("done with bypart blanket")
        else:
            print("No attachments with Component IDs greater than 0.")
            
            
    def getlinks(self):
       # print(self.attachment_dict)
        for attachment,count in self.attachment_dict.items():
            self.list_of_duplicate_byparts.append(attachment)
       # print(self.list_of_duplicate_byparts)
    
    def clear_excel_file(self,file_path):
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None
        wb.save(file_path)
if __name__ == "__main__":
    print("ENTER THE VALID INPUT:")
    print("**********************")
    print(" MENU :")
    print("1: validate")
    print("2: Validate and download")
    print("***********************")
    choice = int(input("ENTER YOUR CHOICE :"))
    db_connection = DatabaseConnection(host="localhost", user="root", password="", database="icm_db")
    db_connection.connect()
    obj = Connect(db_connection.conn)
    #print("entered")

    def choose(choice):
        if choice == 1:
            validate()
        elif choice == 2:
            validate()
            download()
        else:
            print("Invalid choice")

    def validate():
        obj.process_attachments()
        obj.duplicate_byparts()
        obj.getlinks()
        obj.count_of_linktypes()
        obj.partseries()
        obj.bypartandblanket()

    def download():
        obj.download_all_file()

    choose(choice)
    db_connection.close()
