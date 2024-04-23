import mysql.connector
import requests
from collections import Counter
from pathlib import Path
import pandas as pd
import openpyxl
import logging


logging.basicConfig(filename='logfile.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s',filemode='w')

class DatabaseConnection:
    def __init__(self, host, user, password, database):
        self.host = host
        self.user = user
        self.password = password
        self.database = database
        self.conn = None

    def connect(self):
        try:
            self.conn = mysql.connector.connect(
                host=self.host,
                user=self.user,
                password=self.password,
                database=self.database
            )
            print("Connection established successfully!")
            logging.info("Coonection established successfully")
        except mysql.connector.Error as err:
            print(f"Error: {err}")
            logging.exception("connection failed")

    def close(self):
        if self.conn:
            self.conn.close()
            logging.info("connection closed")

class Connect:
    def __init__(self, conn):
        self.conn = conn
        self.cursor = self.conn.cursor()
        self.link = []
        self.path = []
        self.attachment_dict = {}
        self.list_of_duplicate_byparts = []
        self.m_list=[]

    def process_attachments(self):
        query = "SELECT * FROM attachments"
        self.cursor.execute(query)
        self.result = self.cursor.fetchall()
        self.column_index = self.cursor.column_names.index('attachment_filename')
        self.attachments_type = self.cursor.column_names.index('approval_type')
        
        
        query_by_part = "SELECT * FROM attachments WHERE approval_type = 'By part'"
        self.cursor.execute(query_by_part)
        self.result_by_part = self.cursor.fetchall()
        self.result_attachmentname = self.cursor.column_names.index("attachment_filename")
        self.result_component_id=self.cursor.column_names.index("component_id")
        
        
        query1 = "SELECT * FROM manufacturers"
        self.cursor.execute(query1)
        self.manufacture_results = self.cursor.fetchall()
        #column_names.index is a method to access the index of the particular colum as the list item cant be accessed by its 
        self.manufacture_name_manufacture = self.cursor.column_names.index('name')
        self.manufucture_index = self.cursor.column_names.index('normalized_name')
        self.manufacturer_id_index = self.cursor.column_names.index('id')

        # Query to execute the master_components table
        query2 = "SELECT * FROM master_components"
        self.cursor.execute(query2)
        self.component_results = self.cursor.fetchall()
        self.manufacture_part_in_master =self.cursor.column_names.index('manufacturer_partnumber')
        self.component_id_component = self.cursor.column_names.index('id')
        self.manufacture_name_component =self.cursor.column_names.index('manufacturer_name')

        for row in self.result:
            if row[self.column_index].startswith("https://") or row[self.column_index].startswith("http://"):
                self.link.append(row[self.column_index])
            else:
                self.path.append(row[self.column_index])

        self.count_of_total_links = len(self.link)



    def download_all_file(self):
        print("entered download_all_files")
        pat = input("Enter the path you want to insert the folders:") 

        replace_choice = input("Do you want to replace existing files? (yes/no): ").strip().lower()
        if replace_choice not in ("yes", "no",'y','n','Y','N','YES','NO'):
            logging.debug("invalid choice for replacing the file")
            #print("Invalid choice. Exiting...")
            return
        try:
            logging.debug(self.list_of_duplicate_byparts)
            #print("these ar ethe files of the duplicate by parts:",self.list_of_duplicate_byparts)
            for row in self.list_of_duplicate_byparts:
                q="select *  from attachments where attachment_filename=%s and approval_type='By part' group by manufacturer_id "
                self.cursor.execute(q,(row,))
                res=self.cursor.fetchall()
                a_id=self.cursor.column_names.index("id")
                afn=self.cursor.column_names.index("attachment_filename")
                a_mid=self.cursor.column_names.index("manufacturer_id")
                #print(res)
                #print("dem")
                if len(res)>1:
                    for m in res:
                        id_for_attachment=m[a_id]
                        dem=m[afn]
                        attachment_mid=m[a_mid]
                        que="select * from attachments where attachment_filename=%s and manufacturer_id=%s and approval_type='By part'"
                        self.cursor.execute(que,(dem,attachment_mid))
                        res_attachment=self.cursor.fetchall()
                        for i in res_attachment:
                            dir_name_multiple = "compilance"
                            
                            local_directory_multiple = Path(pat) / dir_name_multiple
                            if local_directory_multiple.exists():
                                print("already present multiple fiolder")
                                if(replace_choice=="Y" or  replace_choice=="YES" or replace_choice=="yes" or replace_choice=="y"):
                                    self.download_file(dem,local_directory_multiple,id_for_attachment)

                            else:
                                local_directory_multiple.mkdir()
                                print("Directory for multiple files was created :", local_directory_multiple)
                                self.download_file(dem, local_directory_multiple,id_for_attachment)
                            self.m_list.append(dem)

            logging.debug("list of attachment_files that were mapped to the differnt manufactures with approval type as bypart: %s",self.m_list)
            #print("M_LIST IS:",self.m_list)                

            for row in self.manufacture_results:

                manufacturer_id = row[self.manufacturer_id_index]
                directory_name = row[self.manufucture_index]
                manu_name = row[self.manufacture_name_manufacture]
                local_directory = Path(pat) / directory_name

                if local_directory.exists():
                    logging.debug("already exists:%s", local_directory)
                    #print("already exists:", local_directory)
                else:
                    local_directory.mkdir()
                    logging.debug("Directory created:%s", local_directory)
                   # print("Directory created:", local_directory)

                directory_name_comp = 'compilance'
                local_directory_compilance = local_directory / directory_name_comp

                if local_directory_compilance.exists():
                    logging.debug("already present compilance directory")
                   # print("already present compilance directory")
                else:
                    local_directory_compilance.mkdir()
                    logging.debug("Compilance directory created: %s", local_directory_compilance)
                    #print("Compilance directory created:", local_directory_compilance)

                query2 = "SELECT * FROM master_components WHERE manufacturer_name = %s"
                self.cursor.execute(query2, (manu_name,))
                component_results = self.cursor.fetchall()

                manufacture_part_in_master = self.cursor.column_names.index('manufacturer_partnumber')
                component_id_component = self.cursor.column_names.index('id')
                manufacture_name_component = self.cursor.column_names.index('manufacturer_name')

                for component_row in component_results:
                    if component_row[manufacture_name_component] == row[self.manufacture_name_manufacture]:
                        component_id = component_row[component_id_component]
                        directory_name1 = component_row[manufacture_part_in_master]
                        local_directory2 = local_directory_compilance / directory_name1

                        if local_directory2.exists():
                            logging.debug("Entered component part folder already exists")
                            #print("Entered component part folder already exists")
                        else:
                            local_directory2.mkdir(parents=True, exist_ok=True)
                            logging.debug("Subdirectory created: %s",local_directory2)
                            #print(f"Subdirectory created: {local_directory2}")

                        if self.list_of_duplicate_byparts:
                            attachment_placeholders = ', '.join(['%s'] * len(self.list_of_duplicate_byparts))
                            query3 = f"SELECT * FROM attachments WHERE manufacturer_id = %s AND component_id = %s AND attachment_filename NOT IN ({attachment_placeholders})"
                            query_params = [manufacturer_id, component_id] + self.list_of_duplicate_byparts
                            self.cursor.execute(query3, query_params)
                            attachment_results = self.cursor.fetchall()
                            id_of_index = self.cursor.column_names.index('id')

                            att_res_lin = self.cursor.column_names.index("attachment_filename")

                            #print("attachment result is:",attachment_results)

                            for attachment_row in attachment_results:
                                file_url = attachment_row[att_res_lin]
                                file_type = attachment_row[self.attachments_type]
                                id_of_a = attachment_row[id_of_index]

                                if file_type == "partseries":
                                    directory_name_partseries = 'partseries'
                                    local_directory_partseries = local_directory_compilance / directory_name_partseries
                                    if local_directory_partseries.exists():
                                        logging.debug("already exists : %s",local_directory_partseries)
                                        #print(local_directory_partseries,"already exists")
                                        na=Path(file_url)
                                        last_name=na.name
                                        file_name=Path(local_directory_partseries)/last_name
                                        if(file_name.exists()):
                                            logging.debug("file is already present")
                                            #print("already present:")    
                                            if(replace_choice=='y' or replace_choice=='yes'):
                                                self.download_file(file_url, local_directory_multiple, id_of_a)
                                            else:
                                                local_directory_partseries.mkdir()
                                                logging.debug("Partseries directory created: %s", local_directory_partseries)
                                                #print("Partseries directory created:", local_directory_partseries)
                                                self.download_file(file_url, local_directory_partseries, id_of_a)
                                elif file_type == "blanket":
                                    directory_name_blanket = 'blanket'
                                    local_directory_blanket = local_directory_compilance / directory_name_blanket

                                    if local_directory_blanket.exists():
                                        logging.debug("directory for the blanket already exist: %s",local_directory_blanket)
                                        #print(local_directory_blanket,"already exists")
                                        na=Path(file_url)
                                        last_name=na.name
                                        file_name=Path(local_directory_blanket)/last_name
                                        if(file_name.exists()):
                                            logging.debug("blanket file already present")
                                           # print("already present:")    
                                            if(replace_choice=='y' or replace_choice=='yes'):
                                                self.download_file(file_url, local_directory_multiple, id_of_a)
                                            else:
                                                local_directory_blanket.mkdir()
                                                logging.debug("partseries directory created:%s", local_directory_blanket)
                                                #print("Partseries directory created:", local_directory_blanket)
                                                self.download_file(file_url, local_directory_blanket, id_of_a)

                                else:
                                    na=Path(file_url)
                                    last_name=na.name
                                    file_name=Path(local_directory2)/last_name
                                    #print("local dir:",local_directory2)
                                    #print("By part filame is:",file_name)
                                   # print("id of attachment is :",id_of_a)
                                    if(file_name.exists()):
                                        logging.debug("file already present in:%s",local_directory2)
                                        #print("file already present in:",local_directory2)
                                        if(replace_choice=='y' or replace_choice=='yes'):
                                                #print("entered bro in the folder")
                                                #print("dem is:",file_url)
                                                self.download_file(file_url, local_directory2, id_of_a)
                                    else:
                                        #print("bloody")
                                        #print("filename",file_name)
                                        #print(id_of_a)
                                        self.download_file(file_url, local_directory2,id_of_a)
            self.m_list = set(self.m_list)
            logging.debug("these are the files that are mapped to the same different manufacturers and these are the bypart files:%s",self.m_list)
            #print("these are the files that are mapped to the same different manufacturers and these are the bypart files:",self.m_list)
           # print(self.m_list)
            for k in self.m_list:
                self.list_of_duplicate_byparts.remove(k)
            #print("these are the bypart files mapped to same manfacturer but to the differwnt components",self.list_of_duplicate_byparts)
            for i in self.list_of_duplicate_byparts:
                qp = "SELECT * FROM attachments WHERE approval_type='By part' AND attachment_filename=%s"
                self.cursor.execute(qp, (i,))
                m_id = self.cursor.column_names.index("manufacturer_id")
                ans = self.cursor.fetchall()
                id_in_index = self.cursor.column_names.index("id")

                for j in ans:
                    id_in_attachments = j[id_in_index]
                    aq = "SELECT * FROM manufacturers WHERE id=%s"
                    self.cursor.execute(aq, (j[m_id],))
                    ans_in_j = self.cursor.fetchall()
                    n = self.cursor.column_names.index("normalized_name")
                    for k in ans_in_j:
                        n_name = k[n]
                        dir = Path(pat) / n_name / "compilance" / "multiplebyparts"
                        if dir.exists():
                            logging.debug("already multiple exists")
                            print("already multiple exists")
                            if(replace_choice=='y' or replace_choice=='yes' or replace_choice=="YES" or replace_choice=="Y"):
                                self.download_file(i,dir,id_in_attachments)
                        else:
                            dir.mkdir(parents=True, exist_ok=True)
                            logging.info("created")
                            #print("created")
                            self.download_file(i, dir, id_in_attachments)
        except Exception as e:
            logging.error(e)
            print(f"Error: {e}")

            
    def download_file(self, file_url, local_directory, id_of_attachments):
        try:
            #print("error in tjhis")
            # Download the file from the URL
            response = requests.get(file_url)
            filename = local_directory / Path(file_url.split('?')[0]).name.strip()
            #print(filename)
            # if filename.exists():
            #     print("File already exists:")
            #     return True
            with open(filename, 'wb') as local_file:
                local_file.write(response.content)
                #print("x")
                logging.info("File downloaded and saved as:%s",filename)
                print(f"File downloaded and saved as {filename}")
            #print("completed")
            # Manipulate file path to get the relative path from 'atom_mfr_docs'
            pa = Path(local_directory) / filename
            part = pa.parts
            index = part.index("atom_mfr_docs")
            list_of_parts = list(part[index:])
            finalpath = "\\".join(list_of_parts)
            #print(id_of_attachments)

            # Update the database record with the path of the downloaded file
            query_to_update = "UPDATE attachments SET notes=%s WHERE id=%s"
            self.cursor.execute(query_to_update, (finalpath, id_of_attachments))
            logging.debug("File path updated in the database:%s", finalpath)
            print("File path updated in the database:", finalpath)
            #print("before")
            # Commit the transaction
            self.conn.commit()
            
            return True  # Return True to indicate successful download and database update

        except Exception as e:
            print("An exception occurred:", e)
            return False  # Return False to indicate failure


    def count_of_linktypes(self):
        dict_count_links = {}
        for link_item in self.link:
            count_bypart = 0
            count_partseries = 0
            for row in self.result:
                if row[self.column_index] == link_item:
                    if row[self.attachments_type] == 'By part': 
                        count_bypart += 1
                    elif row[self.attachments_type] == 'partseries':
                        count_partseries += 1
            dict_count_links[link_item] = {'By part': count_bypart, 'Partseries': count_partseries}

        link_count = Counter(self.link)
        for j, count in link_count.items():
            dict_count_links[j] = count

        df_link_count = pd.DataFrame(dict_count_links.items(), columns=['Link', 'Count'])
        df_link_types = pd.DataFrame.from_dict(dict_count_links, orient='index').reset_index()

        with pd.ExcelWriter('count_of_linktypes.xlsx') as writer:
            df_link_count.to_excel(writer, sheet_name='Link Counts', index=False)
            df_link_types.to_excel(writer, sheet_name='Link Details', index=False)

    def duplicate_byparts(self):
        self.attachment_dict = {}
        self.count_of_duplicatebyparts=0
        for link in self.link:
            for row in self.result:
                if row[self.attachments_type] == 'By part' and row[self.column_index]==link:
                    attachment_filename=row[self.column_index]
                    query_component_id = "SELECT component_id FROM attachments WHERE attachment_filename = %s AND approval_type = 'By part'"
                    self.cursor.execute(query_component_id, (attachment_filename,))
                    result_component_ids = self.cursor.fetchall()
                
                     # Extract component IDs
                    component_ids = [res[0] for res in result_component_ids]

                # Check if there are duplicate component IDs
                    if len(set(component_ids)) > 1:
                        self.attachment_dict[attachment_filename] = component_ids
        for attachment,component_id in self.attachment_dict.items():
            self.attachment_dict[attachment]=set(component_id)
        self.count_of_duplicatebyparts=len(self.attachment_dict)

        if self.attachment_dict:
            df_attachment_dict = pd.DataFrame(self.attachment_dict.items(), columns=['Attachment', 'Component IDs'])
            df_duplicate_bypart = pd.DataFrame([['duplicate by parts', self.count_of_duplicatebyparts]],
                                               columns=['Attachment', 'Count'])
            df_total_links = pd.DataFrame([['total links', self.count_of_total_links]],
                                          columns=['Attachment', 'Count'])
            count_of_uniquelinks = len(set(self.link))
            df_uniquelinks = pd.DataFrame([['unique links', count_of_uniquelinks]],
                                          columns=['Attachment', 'Count'])

            with pd.ExcelWriter('duplicate_byparts.xlsx') as writer:
                df_attachment_dict.to_excel(writer, index=False)

            filename = Path(r'C:\Users\mohanvamsi.sadhu\Desktop\dem\total.xlsx')
            if filename.exists():
                self.clear_excel_file(filename)
                wb = openpyxl.load_workbook('total.xlsx')
                ws = wb.active

                ws.append(['count of total links', len(self.link)])
                ws.append(['count of unique links', count_of_uniquelinks])
                ws.append(['count of duplicate by parts', self.count_of_duplicatebyparts])
                wb.save('total.xlsx')

            else:
                with pd.ExcelWriter('total.xlsx') as writer:
                    df_total_links.to_excel(writer, index=False)
                wb = openpyxl.load_workbook('total.xlsx')
                ws = wb.active
                ws.append(['count of unique links ', count_of_uniquelinks])
                ws.append(['count of duplicate by parts', self.count_of_duplicatebyparts])
                wb.save('total.xlsx')
                
                
    def partseries(self):
        self.count_of_partseries_bypart=0
        print("partseries")

        
        # Query to select attachments with approval type 'partseries'
        query_part_series = "SELECT * FROM attachments WHERE approval_type = 'partseries'"
        self.cursor.execute(query_part_series)
        result_part_series = self.cursor.fetchall()
        comp_id_q3 = self.cursor.column_names.index("component_id")
        attach_filename_q3 = self.cursor.column_names.index("attachment_filename")
        
        attachment_dict = {}
        for row1 in self.result_by_part:
            component_ids = []
            for row in result_part_series:
                if row[attach_filename_q3] == row1[self.result_attachmentname]:
                    component_ids.append(row1[self.result_component_id])
                    component_ids.append(row[comp_id_q3])           
            attachment_dict[row1[self.result_attachmentname]] =component_ids
        #print(attachment_dict)

        attachment_dict_filtered = {attachment: component_ids 
                                    for attachment, component_ids in attachment_dict.items() 
                                    if set(component_ids) and len(component_ids) > 0  }
        #print(attachment_dict_filtered)
        
        if attachment_dict_filtered:
            print("Attachments with Component IDs:")
            for attachment, component_ids in attachment_dict_filtered.items():
                if(len(component_ids)>0):
                    print(f"Attachment: {attachment}")
                    print("Component IDs:")
                    for component_id in set(component_ids):
                        print(component_id)
            count_of_parttseries_bypart=len(attachment_dict_filtered)
            # create a excel with the column names as the attachments and the component ID'S
            df_attachment_dict = pd.DataFrame(attachment_dict_filtered.items(), columns=['Attachment', 'Compnoent IDs'])
            with pd.ExcelWriter('partseriesbypart.xlsx') as writer:
                df_attachment_dict.to_excel(writer, index=False)
            
            wb = openpyxl.load_workbook('total.xlsx')
            ws = wb.active
            ws.append(['count of partseries and bypart ',count_of_parttseries_bypart])
            wb.save('total.xlsx')
        
        else:
            print("No attachments with Component IDs greater than 0.")
    def bypartandblanket(self):
        print("bypart and blanket")
        count_of_blanketandbypart=0
        # Query to select attachments with approval type 'partseries'
        query_part_series = "SELECT * FROM attachments WHERE approval_type = 'blanket'"
        self.cursor.execute(query_part_series)
        result_part_series = self.cursor.fetchall()
        comp_id_q3 = self.cursor.column_names.index("component_id")
        attach_filename_q3 = self.cursor.column_names.index("attachment_filename")
        attachment_dict = {}
        for row1 in self.result_by_part:
            component_ids = []
            for row in result_part_series:
                if row[attach_filename_q3] == row1[self.result_attachmentname]:
                    component_ids.append(row[comp_id_q3])   
                    component_ids.append(row1[self.result_component_id])        
            attachment_dict[row1[self.result_attachmentname]] = component_ids
    
        # Filter attachments with component IDs greater than 0
        attachment_dict_filtered = {attachment: component_ids 
                                    for attachment, component_ids in attachment_dict.items() 
                                    if set(component_ids) and len(component_ids) > 0  }
        print(attachment_dict_filtered)
        if attachment_dict_filtered:
            print("Attachments with Component IDs:")
            for attachment, component_ids in attachment_dict_filtered.items():
                print(f"Attachment: {attachment}")
                print("Component IDs:")
                for component_id in set(component_ids):
                    print(component_id)
            for attachment,component_ids in attachment_dict_filtered.items():
                attachment_dict_filtered[attachment]=set(component_ids)
                    
            count_of_blanketandbypart=len(attachment_dict_filtered)
            df_attachment_dict = pd.DataFrame(attachment_dict_filtered.items(), columns=['Attachment', 'Compnoent IDs'])
            df_counts=pd.DataFrame([['blanket and partseries', count_of_blanketandbypart]], columns=['Attachment', 'Count'])
            # Save DataFrame to Excel file 'duplicate_byparts.xlsx'
            with pd.ExcelWriter('bypartblanket.xlsx') as writer:
                df_attachment_dict.to_excel(writer, index=False)
                
            wb = openpyxl.load_workbook('total.xlsx')
            ws = wb.active
            ws.append(['count of partseries and bypart ',count_of_blanketandbypart])
            wb.save('total.xlsx')
            
            print("done with bypart blanket")
        else:
            print("No attachments with Component IDs greater than 0.")
            
            
    def getlinks(self):
       # print(self.attachment_dict)
        for attachment,count in self.attachment_dict.items():
            self.list_of_duplicate_byparts.append(attachment)
       # print(self.list_of_duplicate_byparts)
    
    def clear_excel_file(self,file_path):
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None
        wb.save(file_path)
if __name__ == "__main__":
    print("ENTER THE VALID INPUT:")
    print("**********************")
    print(" MENU :")
    print("1: validate")
    print("2: Validate and download")
    print("***********************")
    choice = int(input("ENTER YOUR CHOICE :"))
    db_connection = DatabaseConnection(host="localhost", user="root", password="", database="icm_db")
    db_connection.connect()
    obj = Connect(db_connection.conn)
    # print("entered")

    def choose(choice):
        if choice == 1:
            validate()
        elif choice == 2:
            validate()
            download()
        else:
            print("Invalid choice")

    def validate():
        obj.process_attachments()
        obj.duplicate_byparts()
        obj.getlinks()
        obj.count_of_linktypes()
        obj.partseries()
        obj.bypartandblanket()

    def download():
        obj.download_all_file()

    choose(choice)
    db_connection.close()
